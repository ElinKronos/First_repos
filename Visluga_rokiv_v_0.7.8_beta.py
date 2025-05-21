
import sys, os, sqlite3, openpyxl, openpyxl.styles, calendar
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
    QLabel, QDialog, QLineEdit, QComboBox, QFrame, QFileDialog, QCheckBox, QMessageBox, QHeaderView, QAbstractItemView
)
from PyQt6.QtGui import QAction
from PyQt6.QtCore import Qt, QTimer
from datetime import datetime, date
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def init_db(db_name=None):
    db_path = os.path.join(os.getcwd(), db_name)
    conn = sqlite3.connect(db_path)
    curs = conn.cursor()

    curs.execute("""
        CREATE TABLE IF NOT EXISTS people (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rank TEXT NOT NULL,
            sec_name TEXT NOT NULL,
            name TEXT NOT NULL,
            unit TEXT NOT NULL,
            note TEXT
        )
    """)

    curs.execute("""
        CREATE TABLE IF NOT EXISTS service_periods (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            person_id INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            FOREIGN KEY (person_id) REFERENCES people(id) ON DELETE CASCADE
        )
    """)

    curs.execute("""
        CREATE TABLE IF NOT EXISTS preferenced_periods (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            person_id INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            preference_type TEXT NOT NULL,
            FOREIGN KEY (person_id) REFERENCES people(id) ON DELETE CASCADE       
        )
    """)

    curs.execute("""
        CREATE TABLE IF NOT EXISTS service_totals (
            person_id INTEGER PRIMARY KEY,
            calendar_years TEXT,
            preferenced_years TEXT,
            FOREIGN KEY (person_id) REFERENCES people(id) ON DELETE CASCADE
        )
    """)

    conn.commit()
    conn.close()
    print("Базу даних ініціалізовано.")

class MainProg(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Вислуга років")
        self.setGeometry(100, 100, 1400, 750)
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.current_db = None

        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu("Файл")
        help_menu = menu_bar.addMenu("Довідка")

        create_db_action = QAction("Створити нову базу даних", self)
        create_db_action.triggered.connect(self.create_new_database)
        file_menu.addAction(create_db_action)

        open_db_action = QAction("Відкрити базу даних", self)
        open_db_action.triggered.connect(self.open_existing_database)
        file_menu.addAction(open_db_action)

        export_to_excel_action = QAction("Експортувати у Excel", self)
        export_to_excel_action.triggered.connect(self.export_to_excel)
        file_menu.addAction(export_to_excel_action)

        exit_action = QAction("Вийти", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        about_action = QAction("Про програму", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

        self.main_layout = QVBoxLayout()
        self.central_widget.setLayout(self.main_layout)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Пошук...")
        self.search_input.textChanged.connect(self.filter_infos)
        
        butt_size = 150

        self.add_people_button = QPushButton("Додати")
        self.add_people_button.setFixedWidth(butt_size)
        self.add_people_button.clicked.connect(self.open_add_people_dialog)

        self.edit_people_button = QPushButton("Редагувати")
        self.edit_people_button.setFixedWidth(butt_size)
        self.edit_people_button.clicked.connect(self.edit_selected_people)

        self.add_period_cal_button = QPushButton("Додати календарі")
        self.add_period_cal_button.setFixedWidth(butt_size)
        self.add_period_cal_button.clicked.connect(self.open_add_period_cal_dialog)

        self.add_period_pref_button = QPushButton("Додати пільгу")
        self.add_period_pref_button.setFixedWidth(butt_size)
        self.add_period_pref_button.clicked.connect(self.open_add_period_pref_dialog)

        self.edit_periods_button = QPushButton("Редагувати вислугу")
        self.edit_periods_button.setFixedWidth(butt_size)
        self.edit_periods_button.clicked.connect(self.open_edit_periods_dialog)

        self.del_people_button = QPushButton("Видалити")
        self.del_people_button.setFixedWidth(butt_size)
        self.del_people_button.clicked.connect(self.del_selected_people)
                        
        top_main_layout = QHBoxLayout()
        top_main_layout.addWidget(self.search_input)
        top_main_layout.addWidget(self.add_people_button)
        top_main_layout.addWidget(self.edit_people_button)        
        top_main_layout.addWidget(self.add_period_cal_button)
        top_main_layout.addWidget(self.add_period_pref_button)
        top_main_layout.addWidget(self.edit_periods_button)
        top_main_layout.addWidget(self.del_people_button)
        self.main_layout.addLayout(top_main_layout)

        self.headers = ["№", "Військове звання", "Прізвище", "Ім'я\nПо батькові", "Підрозділ",
                        "Календарна вислуга років", "Пільгова вислуга років", "Періоди військової служби",
                        "Пільгові періоди служби", "Навчання в цивільному ВНЗ", "Примітка"]

        self.table = QTableWidget()
        self.table.setSortingEnabled(True)
        self.table.setWordWrap(True)
        self.table.resizeRowsToContents()
        self.table.verticalHeader().setVisible(False)
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)
        self.table.horizontalHeader().setStyleSheet("""
            QHeaderView:section {
                background: white;
                border: none;
                border-bottom: 1px solid black;
                border-left: none;
                border-right: 1px solid #cccccc;
                padding: 0px;
                font-weight: bold;
                text-align: center;
            }
        """)

        for col in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col)
            if header_item:
                header_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                font = header_item.font()
                font.setBold(True)
                header_item.setFont(font)

        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.main_layout.addWidget(self.table)
                
        self.infos = []

        last_db_path = os.path.join(os.getcwd(), "last_db.txt")
        if os.path.exists(last_db_path):
            with open(last_db_path, "r", encoding="utf-8") as f:
                self.current_db = f.read().strip()
            if os.path.exists(self.current_db):
                self.load_people_from_db()
                self.load_service_periods()
                self.load_preference_periods()
                self.calculate_totals()
                self.update_table()

        QTimer.singleShot(100, self.show_welcome_message)

    def create_new_database(self):
        dialog = CreateDatabaseDialog(self)
        if dialog.exec():
            db_name = dialog.get_database_name()
            if db_name:
                self.current_db = db_name
                init_db(self.current_db)
                self.load_people_from_db()
                self.load_service_periods()
                self.load_preference_periods()
                self.calculate_totals()
                with open("last_db.txt", "w", encoding="utf-8") as f:
                    f.write(self.current_db)
                QMessageBox.information(self, "Успіх", f"Базу даних '{db_name}' створено успішно")

    def open_existing_database(self):
        file_dialog = QFileDialog(self)
        file_dialog.setNameFilter("Database Files (*.db)")
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFile)

        if file_dialog.exec():
            selected_file = file_dialog.selectedFiles()[0]
            if selected_file:
                self.current_db = selected_file
                self.load_people_from_db()
                self.load_service_periods()
                self.load_preference_periods()
                self.calculate_totals()
                with open("last_db.txt", "w", encoding="utf-8") as f:
                    f.write(self.current_db)
                QMessageBox.information(self, "Успіх", f"Базу даних '{os.path.basename(selected_file)}' успішно відкрито")

    def open_add_people_dialog(self):
        dialog = AddPeople(self, info = {"rank": "", "sec_name": "", "name": "", "unit": "", "note": ""})
        if dialog.exec():
            info_data_people = dialog.get_info_data_people()
            self.add_people(info_data_people)

    def open_add_period_cal_dialog(self):
        selected = self.table.currentRow()
        if selected < 0 or selected >= len(self.infos):
            return
        
        person_id = self.infos[selected]["id"]
        person_data = self.infos[selected]
        dialog = AddPeriod_Calendar(self, person_data, person_id=person_id,)
        if dialog.exec():
            info_data_period_cal = dialog.get_info_data_period_cal()
            self.add_period_cal(info_data_period_cal)
    
    def open_edit_periods_dialog(self):
        selected = self.table.currentRow()
        if selected < 0 or selected >= len(self.infos):
            return
        
        person_data = self.infos[selected]
        dialog = EditPeriodsDialog(self, person_data, self.current_db)
        dialog.exec()
        self.load_service_periods()
        self.load_preference_periods()
        self.calculate_totals()
        self.update_table()

    def add_period_cal(self, period_data):
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()
        curs.execute("""
            INSERT INTO service_periods (person_id, start_date, end_date)
            VALUES (?, ?, ?)
        """, (
            period_data["person_id"],
            period_data["start_date"],
            period_data["end_date"]
        ))
        conn.commit()
        conn.close()
        self.load_service_periods()
        self.update_table()
        self.calculate_totals()

    def open_add_period_pref_dialog(self):
        selected = self.table.currentRow()
        if selected < 0 or selected >= len(self.infos):
            return
            
        person_id = self.infos[selected]["id"]
        person_data = self.infos[selected]
        dialog = AddPeriod_Pref(self, person_data, person_id=person_id)
        if dialog.exec():
            info_data_period_pref = dialog.get_info_data_period_pref()
            self.add_period_pref(info_data_period_pref)
       
    def add_period_pref(self, period_data):
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()
        curs.execute("""
            INSERT INTO preferenced_periods (person_id, start_date, end_date, preference_type)
            VALUES (?, ?, ?, ?)
        """, (
            period_data["person_id"],
            period_data["start_date"],
            period_data["end_date"],
            period_data["preference_type"]
        ))
        conn.commit()
        conn.close()
        self.load_preference_periods()
        self.update_table()
        self.calculate_totals()

    def add_people(self, info_data):
        self.add_people_to_db(info_data)
        self.infos.append(info_data)
        self.update_table()

    def add_people_to_db(self, info_data):
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()

        curs.execute("""
            INSERT INTO people (rank, sec_name, name, unit, note)
            VALUES (?, ?, ?, ?, ?)
        """, (
            info_data["rank"],
            info_data["sec_name"],
            info_data["name"],
            info_data["unit"],
            info_data["note"]
        ))

        info_data["id"] = curs.lastrowid

        conn.commit()
        conn.close()

    def load_people_from_db(self):
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()

        curs.execute("SELECT id, rank, sec_name, name, unit, note FROM people")
        rows = curs.fetchall()

        self.infos = [
            {"id": row[0], "rank": row[1], "sec_name": row[2], "name": row[3], "unit": row[4], "note": row[5]}
            for row in rows
        ]

        conn.close()
        self.update_table()

    def load_service_periods(self):
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()

        for info in self.infos:
            curs.execute("""
                SELECT start_date, end_date FROM service_periods
                WHERE person_id = (
                    SELECT id FROM people
                    WHERE rank=? AND sec_name=? AND name=? AND unit=? AND note=?         
                )
            """, (
                info["rank"], info["sec_name"], info["name"], info["unit"], info["note"]
            ))

            periods = curs.fetchall()
            formatted = [f"{self.format_date(p[0])} - {self.format_date(p[1])}" for p in periods]
            info["cal_periods"] = "\n".join(formatted) if formatted else ""

        conn.close()

    def load_preference_periods(self):
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()

        for info in self.infos:
            curs.execute("""
                SELECT start_date, end_date, preference_type FROM preferenced_periods
                WHERE person_id = (
                    SELECT id FROM people
                    WHERE rank=? AND sec_name=? AND name=? AND unit=? AND note=?
                )
            """, (
                info["rank"], info["sec_name"], info["name"], info["unit"], info["note"]
            ))

            periods = curs.fetchall()
            general = []
            civil = []

            for start, end, pref_type in periods:
                if pref_type.strip().lower() == "навчання у цивільному внз":
                    formatted = f"{self.format_date(start)} - {self.format_date(end)}"
                    civil.append(formatted)
                else:
                    formatted = f"{self.format_date(start)} - {self.format_date(end)} ({pref_type})"
                    general.append(formatted)

            info["pref_periods"] = "\n".join(general) if general else ""
            info["civil_edu"] = "\n" .join(civil) if civil else ""

        conn.close()

    def edit_people(self, info_id):
        person_data = self.infos[info_id]
        dialog = AddPeople(self, person_data)
        if dialog.exec():
            new_data = dialog.get_info_data_people()
            
            person_id = person_data["id"]
            self.update_people_in_db(person_id, new_data)
            
            new_data["id"] = person_id
            self.infos[info_id] = new_data

            self.load_service_periods()
            self.load_preference_periods()
            self.calculate_totals()
            self.update_table()

    def update_people_in_db(self, person_id, new_data):
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()

        curs.execute("""
            UPDATE people
            SET rank = ?, sec_name = ?, name = ?, unit = ?, note = ?
            WHERE id = ?

        """, (
            new_data["rank"],
            new_data["sec_name"],
            new_data["name"],
            new_data["unit"],
            new_data["note"],
            person_id
        ))
    
        conn.commit()
        conn.close()

    def filter_infos(self):
        search_text = self.search_input.text().lower()
        self.filtered_infos = [info for info in self.infos if search_text in info["sec_name"].lower() 
                               or search_text in info["name"].lower()]
        self.update_table(filtered=True)

    def update_table(self, filtered=False):
        infos_to_display = self.filtered_infos if filtered else self.infos
        self.table.setRowCount(len(infos_to_display))

        for row_idx, info in enumerate(infos_to_display):
            
            item = QTableWidgetItem(str(row_idx + 1))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 0, item)
            
            item = QTableWidgetItem(info.get("rank", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 1, item)
            
            item = QTableWidgetItem(info.get("sec_name", "").upper())
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 2, item)
            
            item = QTableWidgetItem(info.get("name", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 3, item)
            
            item = QTableWidgetItem(info.get("unit", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 4, item)
            
            item = QTableWidgetItem(info.get("cal_SY", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 5, item)
            
            item = QTableWidgetItem(info.get("pref_SY", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 6, item)
            
            item = QTableWidgetItem(info.get("cal_periods", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 7, item)
            
            item = QTableWidgetItem(info.get("pref_periods", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 8, item)
            
            item = QTableWidgetItem(info.get("civil_edu", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 9, item)
            
            item = QTableWidgetItem(info.get("note", ""))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 10, item)
        
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()
        
    def export_database_to_excel(self, output_path):
    
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Вислуга років"

        headers = ["№", "Військове звання", "Прізвище", "Ім'я\nПо батькові", "Підрозділ",
                    "Календарна вислуга років", "Пільгова вислуга років", "Періоди військової служби",
                    "Пільгові періоди служби", "Навчання в цивільному ВНЗ", "Примітка"]
    
        ws.append(headers)

        for index, person in enumerate(self.infos):
            ws.append([
                str(index + 1),
                person.get("rank", ""),
                person.get("sec_name", ""),
                person.get("name", ""),
                person.get("unit", ""),
                person.get("cal_SY", ""),
                person.get("pref_SY", ""),
                person.get("cal_periods", ""),
                person.get("pref_periods", ""),
                person.get("civil_edu", ""),
                person.get("note", "")
            ])

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        wb.save(output_path)

    def export_to_excel(self):
        path, _= QFileDialog.getSaveFileName(
            self,
            "Зберегти Excel-файл",
            "База даних вислуги років_.xlsx",
            "excel файли (*.xlsx)"
        )
        if path:
            self.export_database_to_excel(path)

            QMessageBox.information(
                self,
                "Експортування завершено",
                f"Базу успішно експортовано у файл:\n{path}"
            )

    def edit_selected_people(self):
        selected = self.table.currentRow()
        if selected >= 0 and selected < len(self.infos):
            self.edit_people(selected)

    def show_about(self):
        about = AboutDialog(self)
        about.exec()

    def del_selected_people(self):
        selected = self.table.currentRow()
        if selected < 0 or selected >= len(self.infos):
            return

        msg = QMessageBox(self)
        msg.setWindowTitle("Підтвердження видалення")
        msg.setText("Ви справді хочете видалити цю особу?")
        msg.setIcon(QMessageBox.Icon.Question)
        msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        yes_button = msg.button(QMessageBox.StandardButton.Yes)
        no_button = msg.button(QMessageBox.StandardButton.No)
        yes_button.setText("Так")
        no_button.setText("Ні")

        msg.exec()

        if msg.clickedButton() == yes_button:
            person_id = self.infos[selected]["id"]

            conn = sqlite3.connect(self.current_db)
            curs = conn.cursor()
            curs.execute("DELETE FROM people WHERE id = ?", (person_id,))
            conn.commit()
            conn.close()

            del self.infos[selected]
            self.update_table()

    def format_date(self, date_str):
        if date_str == "NOW":
            return "по т.ч."
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            return "???"
    
    def calculate_exact_period(self, start_date, end_date):
        if end_date < start_date:
            return "0 р. 0 м. 0 д."
        
        sy, sm, sd = start_date.year, start_date.month, start_date.day
        ey, em, ed = end_date.year, end_date.month, end_date.day

        years = ey - sy
        months = em - sm
        days = ed - sd

        if days < 0:
            months -= 1
            days += 30

        if months < 0:
            years -= 1
            months += 12

        return f"{years} р. {months} м. {days} д."

    def date_diff(self, start, end):
        sy, sm, sd = start.year, start.month, start.day
        ey, em, ed = end.year, end.month, end.day
        
        years = ey - sy
        months = em - sm
        days = ed - sd

        if days < 0:
            months -= 1
            days += 30

        if months < 0:
            years -= 1
            months += 12

        return years, months, days
            
    def normalize_ymd(self, y, m, d):
        if d >= 30:
            m += 1
            d -= 30
        if m >= 12:
            y += 1
            m -= 12
        return y, m, d

    def calculate_totals(self):
        if not self.infos:
            return
        
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()

        coeffs = {
            "1 день/3 дні": 3.0,
            "1 день/2 дні": 2.0,
            "1 день/1,5 дні": 1.5,
            "30 днів/40 днів": 1.33,
            "Навчання у цивільному ВНЗ": 0.5
        }
        max_civil = (2, 6, 0)

        for info in self.infos:
            person_id = info["id"]

            cal_years = "0 р. 0 м. 0 д."
            pref_years = "0 р. 0 м. 0 д."

            total_y, total_m, total_d = 0, 0, 0
            curs.execute("SELECT start_date, end_date FROM service_periods WHERE person_id=?", (person_id,))
            for s, e in curs.fetchall():
                start = datetime.strptime(s, "%Y-%m-%d")
                end = datetime.now() if e == "NOW" else datetime.strptime(e, "%Y-%m-%d")
                
                yc, mc, dc = self.date_diff(start, end)
                total_y += yc
                total_m += mc
                total_d += dc

            pref_y, pref_m, pref_d = 0, 0, 0
            civil_y, civil_m, civil_d = 0, 0, 0
            curs.execute("SELECT start_date, end_date, preference_type FROM preferenced_periods WHERE person_id=?", (person_id,))
            for s, e, ptype in curs.fetchall():
                start = datetime.strptime(s, "%Y-%m-%d")
                end = datetime.now() if e == "NOW" else datetime.strptime(e, "%Y-%m-%d")
                
                y, m, d = self.date_diff(start, end)
                coeff = coeffs.get(ptype, 1.0)

                if coeff in [1.5, 1.33, 0.5]:
                    total_days = y * 360 + m * 30 + d
                    converted_days = int(total_days * coeff)
                    py = converted_days // 360
                    pm = (converted_days % 360) // 30
                    pd = (converted_days % 360) % 30
                else:
                    py = int(y * coeff)
                    pm = int(m * coeff)
                    pd = int(d * coeff)

                py_n, pm_n, pd_n = self.normalize_ymd(py, pm, pd)
                
                if ptype == "Навчання у цивільному ВНЗ":
                    civil_y += py_n
                    civil_m += pm_n
                    civil_d += pd_n
                else:
                    pref_y += py_n
                    pref_m += pm_n
                    pref_d += pd_n

            if (civil_y, civil_m, civil_d) > max_civil:
                civil_y, civil_m, civil_d = max_civil

            total_y_c = total_y + civil_y
            total_m_c = total_m + civil_m
            total_d_c = total_d + civil_d
            total_y_c_n, total_m_c_n, total_d_c_n = self.normalize_ymd(total_y_c, total_m_c, total_d_c)

            cal_years = f"{total_y_c_n} р. {total_m_c_n} м. {total_d_c_n} д."

            total_y_p = total_y_c_n + pref_y
            total_m_p = total_m_c_n + pref_m
            total_d_p = total_d_c_n + pref_d
            total_y_p_n, total_m_p_n, total_d_p_n = self.normalize_ymd(total_y_p, total_m_p, total_d_p)

            pref_years = f"{total_y_p_n} р. {total_m_p_n} м. {total_d_p_n} д."

            info["cal_SY"] = cal_years
            info["pref_SY"] = pref_years

            curs.execute("REPLACE INTO service_totals (person_id, calendar_years, preferenced_years) VALUES (?, ?, ?)",
                        (person_id, cal_years, pref_years))

        conn.commit()
        conn.close()
        self.update_table()

    def show_welcome_message(self):
        welcome_text = (
        "<p>Шановні колеги!</p>"
        "<p>Перед вами — бета-версія програми <b>'Вислуга років'</b> версії <b>0.7.8_beta</b>.</p>"
        "<p>Ця система створена з турботою і бажанням допомогти вам у щоденній роботі.<br>"
        "В її розробку вкладено багато часу, праці і частинку душі.</p>"
        "<p>Я буду дуже вдячний за будь-які ваші зауваження, відгуки та побажання.<br>"
        "Кожен ваш відгук допоможе вдосконалити програму та зробити її ще зручнішою для вас.</p>"
        "<p>Дякую, що тестуєте мій проєкт!</p>"
        "<p>З найкращими побажаннями,<br>"
        "✨ <u>Elin & Nyx</u> ✨</p>"
        )

        msg = QMessageBox(self)
        msg.setWindowTitle("Привітання")
        msg.setTextFormat(Qt.TextFormat.RichText)
        msg.setText(welcome_text)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

class AddPeople(QDialog):
    def __init__(self, parent = None, info = None):
        super().__init__(parent)
        self.setWindowTitle("Інформація про військовослужбовця")
        self.setGeometry(200, 200, 300, 300)
        self.setFixedSize(300, 300)

        self.layout = QVBoxLayout()

        self.rank_input = QComboBox()
        self.rank_input.addItems(["полковник", "підполковник", "майор", "капітан", "старший лейтенант",
                                  "лейтенант", "молодший лейтенант", "майстер-сержант", "штаб-сержант",
                                   "головний сержант", "старший сержант", "сержант", "молодший сержант",
                                   "старший матрос", "матрос"])
        if info["rank"] in [self.rank_input.itemText(i) for i in range (self.rank_input.count())]:
            self.rank_input.setCurrentText(info["rank"])
        else:
            self.rank_input.setCurrentText("військове звання")

        self.sec_name_input = QLineEdit(info["sec_name"])

        self.name_input = QLineEdit(info["name"])

        self.unit_input = QComboBox()
        self.unit_input.addItems(["Управління", "1 садн", "2 садн", "3 адн", "4 адн", "Реабатр",
                                  "ДнАР", "Взвод РЕБ", "Рота охорони", "Інженерний взвод", "ПВЗ",
                                  "ІТВ", "Ремонтна рота", "РМЗ", "Пожежний взвод", "КТП", "Оркестр",
                                  "Медичний пункт", "Клуб", "ГКБС", "ГПСВ"])
        if info["unit"] in [self.unit_input.itemText(i) for i in range (self.unit_input.count())]:
            self.unit_input.setCurrentText(info["unit"])
        else:
            self.unit_input.setCurrentText("підрозділ")

        self.note_input = QLineEdit(info["note"])

        self.save_button = QPushButton("Зберегти")
        self.save_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Скасувати")
        self.cancel_button.clicked.connect(self.reject)
        
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(self.cancel_button)
        
        top_layout = QVBoxLayout()
        top_layout.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)
        top_layout.addWidget(QLabel("Військове звання"))
        top_layout.addWidget(self.rank_input)
        top_layout.addWidget(QLabel("Прізвище"))
        top_layout.addWidget(self.sec_name_input)
        top_layout.addWidget(QLabel("Ім'я та по батькові"))
        top_layout.addWidget(self.name_input)
        top_layout.addWidget(QLabel("Підрозділ"))
        top_layout.addWidget(self.unit_input)
        top_layout.addWidget(QLabel("Примітка"))
        top_layout.addWidget(self.note_input)
        
        self.layout.addLayout(top_layout)
        self.layout.addStretch(1)
        self.layout.addLayout(button_layout)

        self.setLayout(self.layout)

    def get_info_data_people(self):
        return {
            "rank": self.rank_input.currentText(),
            "sec_name": self.sec_name_input.text(),
            "name": self.name_input.text(),
            "unit": self.unit_input.currentText(),
            "note": self.note_input.text()
        }

class AddPeriod_Calendar(QDialog):
    def __init__(self, parent, person_data, person_id = None,):
        super().__init__(parent)
        self.person_id = person_id
        self.person_data = person_data
        self.setWindowTitle("Календарна вислуга")
        self.setGeometry(200, 200, 250, 300)
        self.setFixedSize(250, 300)

        self.layout = QVBoxLayout()

        rank = person_data.get("rank", "")
        sec_name = person_data.get("sec_name", "").upper()
        name = person_data.get("name", "")
        
        person_label_rank = QLabel(f"<b>{rank}</b>")
        person_label_rank.setAlignment(Qt.AlignmentFlag.AlignCenter)
        person_label_sec_name = QLabel(f"<b>{sec_name}</b>")
        person_label_sec_name.setAlignment(Qt.AlignmentFlag.AlignCenter)
        person_label_name = QLabel(f"<b>{name}</b>")
        person_label_name.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)

        today = datetime.now()
        current_year = today.year
        current_month = today.month
        current_day = today.day

        self.start_day_input = QComboBox()
        self.start_month_input = QComboBox()
        self.start_year_input = QComboBox()
        self.start_year_input.addItems([str(x) for x in range(datetime.now().year, 1979, -1)])
        
        start_inputs = QHBoxLayout()
        start_inputs.addWidget(self.start_day_input)
        start_inputs.addWidget(self.start_month_input)
        start_inputs.addWidget(self.start_year_input)

        self.end_day_input = QComboBox()
        self.end_month_input = QComboBox()
        self.end_year_input = QComboBox()
        self.end_year_input.addItems([str(x) for x in range(datetime.now().year, 1979, -1)])
        
        end_inputs = QHBoxLayout()
        end_inputs.addWidget(self.end_day_input)
        end_inputs.addWidget(self.end_month_input)
        end_inputs.addWidget(self.end_year_input)

        self.now_checkbox = QCheckBox("по теперішній час")
        self.now_checkbox.stateChanged.connect(self.toggle_end_date_fields)

        self.add_button = QPushButton("Підтвердити")
        self.add_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Вихід")
        self.cancel_button.clicked.connect(self.reject)
        
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.add_button)
        button_layout.addWidget(self.cancel_button)

        top_layout = QVBoxLayout()
        top_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        top_layout.addWidget(person_label_rank)
        top_layout.addWidget(person_label_sec_name)
        top_layout.addWidget(person_label_name)
        top_layout.addWidget(line)
        top_layout.addWidget(QLabel("Дата початку служби"))
        top_layout.addLayout(start_inputs)
        top_layout.addWidget(QLabel("Дата завершення служби"))
        top_layout.addLayout(end_inputs)
        top_layout.addWidget(self.now_checkbox)

        self.layout.addLayout(top_layout)
        self.layout.addLayout(button_layout)

        self.setLayout(self.layout)

        self.start_year_input.currentIndexChanged.connect(self.update_end_constraints)
        self.start_month_input.currentIndexChanged.connect(self.update_end_constraints)
        self.start_day_input.currentIndexChanged.connect(self.update_end_constraints)

        self.end_year_input.currentIndexChanged.connect(self.update_start_constraints)
        self.end_month_input.currentIndexChanged.connect(self.update_start_constraints)
        self.end_day_input.currentIndexChanged.connect(self.update_start_constraints)

        self.start_year_input.currentIndexChanged.connect(lambda: self.update_months(self.start_year_input, self.start_month_input))
        self.start_month_input.currentIndexChanged.connect(lambda: self.update_days(self.start_year_input, self.start_month_input, self.start_day_input))
        self.start_year_input.currentIndexChanged.connect(lambda: self.update_days(self.start_year_input, self.start_month_input, self.start_day_input))

        self.end_year_input.currentIndexChanged.connect(lambda: self.update_months(self.end_year_input, self.end_month_input))
        self.end_month_input.currentIndexChanged.connect(lambda: self.update_days(self.end_year_input, self.end_month_input, self.end_day_input))
        self.end_year_input.currentIndexChanged.connect(lambda: self.update_days(self.end_year_input, self.end_month_input, self.end_day_input))
        
        self.update_months(self.start_year_input, self.start_month_input)
        self.update_days(self.start_year_input, self.start_month_input, self.start_day_input)
        self.update_months(self.end_year_input, self.end_month_input)
        self.update_days(self.end_year_input, self.end_month_input, self.end_day_input)

        self.start_year_input.setCurrentText(str(current_year))
        self.start_month_input.setCurrentText(str(current_month))
        self.start_day_input.setCurrentText(str(current_day))

        self.end_year_input.setCurrentText(str(current_year))
        self.end_month_input.setCurrentText(str(current_month))
        self.end_day_input.setCurrentText(str(current_day))

    def toggle_end_date_fields(self):
        disabled = self.now_checkbox.isChecked()
        self.end_day_input.setDisabled(disabled)
        self.end_month_input.setDisabled(disabled)
        self.end_year_input.setDisabled(disabled)

    def update_months(self, year_input, month_input):
        current = datetime.now()
        year = int(year_input.currentText())
        
        month_input.blockSignals(True)
        month_input.clear()

        if year == current.year:
            month_input.addItems([str(m) for m in range(1, current.month + 1)])
        else:
            month_input.addItems([str(m) for m in range(1, 13)])

        month_input.setCurrentIndex(0)
        month_input.blockSignals(False)

    def update_days(self, year_input, month_input, day_input):
        year = int(year_input.currentText())
        month = int(month_input.currentText())
        current = datetime.now()

        max_day = calendar.monthrange(year, month)[1]
        if year == current.year and month == current.month:
            max_day = min(max_day, current.day)

        day_input.clear()
        day_input.addItems([str(d) for d in range(1, max_day + 1)])

    def update_end_constraints(self):
        try:
            start = date(
                int(self.start_year_input.currentText()),
                int(self.start_month_input.currentText()),
                int(self.start_day_input.currentText())
            )
            end = date(
                int(self.end_year_input.currentText()),
                int(self.end_month_input.currentText()),
                int(self.end_day_input.currentText())
            )
            if end < start:
                self.end_year_input.setCurrentText(self.start_year_input.currentText())
                self.end_month_input.setCurrentText(self.start_month_input.currentText())
                self.end_day_input.setCurrentText(self.start_day_input.currentText())
        except Exception:
            pass

    def update_start_constraints(self):
        try:
            start = date(
                int(self.start_year_input.currentText()),
                int(self.start_month_input.currentText()),
                int(self.start_day_input.currentText())
            )
            end = date(
                int(self.end_year_input.currentText()),
                int(self.end_month_input.currentText()),
                int(self.end_day_input.currentText())
            )
            if start > end:
                self.start_year_input.setCurrentText(self.end_year_input.currentText())
                self.start_month_input.setCurrentText(self.end_month_input.currentText())
                self.start_day_input.setCurrentText(self.end_day_input.currentText())
        except Exception:
            pass

    def get_info_data_period_cal(self):
        start_date = f"{self.start_year_input.currentText()}-{self.start_month_input.currentText().zfill(2)}-{self.start_day_input.currentText().zfill(2)}"
        
        if self.now_checkbox.isChecked():
            end_date = "NOW"
        else:
            end_date = f"{self.end_year_input.currentText()}-{self.end_month_input.currentText().zfill(2)}-{self.end_day_input.currentText().zfill(2)}"
        
        return {
            "person_id": self.person_id,
            "start_date": start_date,
            "end_date": end_date
        }

class AddPeriod_Pref(QDialog):
    def __init__(self, parent, person_data, person_id = None):
        super().__init__(parent)
        self.person_id = person_id
        self.person_data = person_data
        self.setWindowTitle("Пільгова вислуга")
        self.setGeometry(200, 200, 250, 300)
        self.setFixedSize(250, 300)

        self.layout = QVBoxLayout()

        rank = person_data.get("rank", "")
        sec_name = person_data.get("sec_name", "").upper()
        name = person_data.get("name", "")
        
        person_label_rank = QLabel(f"<b>{rank}</b>")
        person_label_rank.setAlignment(Qt.AlignmentFlag.AlignCenter)
        person_label_sec_name = QLabel(f"<b>{sec_name}</b>")
        person_label_sec_name.setAlignment(Qt.AlignmentFlag.AlignCenter)
        person_label_name = QLabel(f"<b>{name}</b>")
        person_label_name.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)

        today = datetime.now()
        current_year = today.year
        current_month = today.month
        current_day = today.day

        self.start_day_input = QComboBox()
        self.start_month_input = QComboBox()
        self.start_year_input = QComboBox()
        self.start_year_input.addItems([str(x) for x in range(datetime.now().year, 1979, -1)])

        start_inputs = QHBoxLayout()
        start_inputs.addWidget(self.start_day_input)
        start_inputs.addWidget(self.start_month_input)
        start_inputs.addWidget(self.start_year_input)

        self.end_day_input = QComboBox()
        self.end_month_input = QComboBox()
        self.end_year_input = QComboBox()
        self.end_year_input.addItems([str(x) for x in range(datetime.now().year, 1979, -1)])
        
        end_inputs = QHBoxLayout()
        end_inputs.addWidget(self.end_day_input)
        end_inputs.addWidget(self.end_month_input)
        end_inputs.addWidget(self.end_year_input)

        self.now_checkbox = QCheckBox("по теперішній час")
        self.now_checkbox.stateChanged.connect(self.toggle_end_date_fields)

        self.pref_type_input = QComboBox()
        self.pref_type_input.addItems(["1 день/3 дні", "1 день/2 дні", "1 день/1,5 дні", "30 днів/40 днів", "Навчання у цивільному ВНЗ"])
        self.pref_type_input.setCurrentIndex(0)

        self.add_button = QPushButton("Підтвердити")
        self.add_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Вихід")
        self.cancel_button.clicked.connect(self.reject)
        
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.add_button)
        button_layout.addWidget(self.cancel_button)

        top_layout = QVBoxLayout()
        top_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        top_layout.addWidget(person_label_rank)
        top_layout.addWidget(person_label_sec_name)
        top_layout.addWidget(person_label_name)
        top_layout.addWidget(line)
        top_layout.addWidget(QLabel("Дата початку пільгового періоду"))
        top_layout.addLayout(start_inputs)
        top_layout.addWidget(QLabel("Дата завершення пільгового періоду"))
        top_layout.addLayout(end_inputs)
        top_layout.addWidget(self.now_checkbox)
        top_layout.addWidget(QLabel("Оберіть тип пільгової вислуги"))
        top_layout.addWidget(self.pref_type_input)

        self.layout.addLayout(top_layout)
        self.layout.addLayout(button_layout)

        self.setLayout(self.layout)

        self.start_year_input.currentIndexChanged.connect(self.update_end_constraints)
        self.start_month_input.currentIndexChanged.connect(self.update_end_constraints)
        self.start_day_input.currentIndexChanged.connect(self.update_end_constraints)

        self.end_year_input.currentIndexChanged.connect(self.update_start_constraints)
        self.end_month_input.currentIndexChanged.connect(self.update_start_constraints)
        self.end_day_input.currentIndexChanged.connect(self.update_start_constraints)

        self.start_year_input.currentIndexChanged.connect(lambda: self.update_months(self.start_year_input, self.start_month_input))
        self.start_month_input.currentIndexChanged.connect(lambda: self.update_days(self.start_year_input, self.start_month_input, self.start_day_input))
        self.start_year_input.currentIndexChanged.connect(lambda: self.update_days(self.start_year_input, self.start_month_input, self.start_day_input))

        self.end_year_input.currentIndexChanged.connect(lambda: self.update_months(self.end_year_input, self.end_month_input))
        self.end_month_input.currentIndexChanged.connect(lambda: self.update_days(self.end_year_input, self.end_month_input, self.end_day_input))
        self.end_year_input.currentIndexChanged.connect(lambda: self.update_days(self.end_year_input, self.end_month_input, self.end_day_input))
        
        self.update_months(self.start_year_input, self.start_month_input)
        self.update_days(self.start_year_input, self.start_month_input, self.start_day_input)
        self.update_months(self.end_year_input, self.end_month_input)
        self.update_days(self.end_year_input, self.end_month_input, self.end_day_input)

        self.start_year_input.setCurrentText(str(current_year))
        self.start_month_input.setCurrentText(str(current_month))
        self.start_day_input.setCurrentText(str(current_day))

        self.end_year_input.setCurrentText(str(current_year))
        self.end_month_input.setCurrentText(str(current_month))
        self.end_day_input.setCurrentText(str(current_day))

    def toggle_end_date_fields(self):
        disabled = self.now_checkbox.isChecked()
        self.end_day_input.setDisabled(disabled)
        self.end_month_input.setDisabled(disabled)
        self.end_year_input.setDisabled(disabled)

    def update_months(self, year_input, month_input):
        current = datetime.now()
        year = int(year_input.currentText())
        
        month_input.blockSignals(True)
        month_input.clear()

        if year == current.year:
            month_input.addItems([str(m) for m in range(1, current.month + 1)])
        else:
            month_input.addItems([str(m) for m in range(1, 13)])

        month_input.setCurrentIndex(0)
        month_input.blockSignals(False)

    def update_days(self, year_input, month_input, day_input):
        year = int(year_input.currentText())
        month = int(month_input.currentText())
        current = datetime.now()

        max_day = calendar.monthrange(year, month)[1]
        if year == current.year and month == current.month:
            max_day = min(max_day, current.day)

        day_input.clear()
        day_input.addItems([str(d) for d in range(1, max_day + 1)])

    def update_end_constraints(self):
        try:
            start = date(
                int(self.start_year_input.currentText()),
                int(self.start_month_input.currentText()),
                int(self.start_day_input.currentText())
            )
            end = date(
                int(self.end_year_input.currentText()),
                int(self.end_month_input.currentText()),
                int(self.end_day_input.currentText())
            )
            if end < start:
                self.end_year_input.setCurrentText(self.start_year_input.currentText())
                self.end_month_input.setCurrentText(self.start_month_input.currentText())
                self.end_day_input.setCurrentText(self.start_day_input.currentText())
        except Exception:
            pass

    def update_start_constraints(self):
        try:
            start = date(
                int(self.start_year_input.currentText()),
                int(self.start_month_input.currentText()),
                int(self.start_day_input.currentText())
            )
            end = date(
                int(self.end_year_input.currentText()),
                int(self.end_month_input.currentText()),
                int(self.end_day_input.currentText())
            )
            if start > end:
                self.start_year_input.setCurrentText(self.end_year_input.currentText())
                self.start_month_input.setCurrentText(self.end_month_input.currentText())
                self.start_day_input.setCurrentText(self.end_day_input.currentText())
        except Exception:
            pass

    def get_info_data_period_pref(self):
        start_date = f"{self.start_year_input.currentText()}-{self.start_month_input.currentText().zfill(2)}-{self.start_day_input.currentText().zfill(2)}"
        
        if self.now_checkbox.isChecked():
            end_date = "NOW"
        else:
            end_date = f"{self.end_year_input.currentText()}-{self.end_month_input.currentText().zfill(2)}-{self.end_day_input.currentText().zfill(2)}"
        
        return {
            "person_id": self.person_id,
            "start_date": start_date,
            "end_date": end_date,
            "preference_type": self.pref_type_input.currentText()
        }

class EditPeriodsDialog(QDialog):
    def __init__(self, parent, person_data, current_db):
        super().__init__(parent)
        self.person_data = person_data
        self.current_db = current_db
        self.setWindowTitle("Редагування періодів служби")
        self.setGeometry(300, 100, 400, 600)
        self.setFixedSize(400, 600)

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        rank = person_data.get("rank", "")
        sec_name = person_data.get("sec_name", "").upper()
        name = person_data.get("name", "")
        
        person_label_rank = QLabel(f"<b>{rank}</b>")
        person_label_rank.setAlignment(Qt.AlignmentFlag.AlignCenter)
        person_label_sec_name = QLabel(f"<b>{sec_name}</b>")
        person_label_sec_name.setAlignment(Qt.AlignmentFlag.AlignCenter)
        person_label_name = QLabel(f"<b>{name}</b>")
        person_label_name.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)

        self.person_id = person_data["id"]

        self.cal_table = QTableWidget(0, 2)
        self.cal_table.setHorizontalHeaderLabels(["Дата початку", "Дата завершення"])
        self.cal_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.cal_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.cal_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        self.pref_table = QTableWidget(0, 3)
        self.pref_table.setHorizontalHeaderLabels(["Дата початку", "Дата завершення", "Тип пільги"])
        self.pref_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.pref_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.pref_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        
        button_layout = QHBoxLayout()
        self.edit_button = QPushButton("Редагувати")
        self.del_button = QPushButton("Видалити")
        self.edit_button.clicked.connect(self.edit_selected_period)
        self.del_button.clicked.connect(self.del_selected_period)
        button_layout.addWidget(self.edit_button)
        button_layout.addWidget(self.del_button)
        
        close_button = QPushButton("Готово")
        close_button.clicked.connect(self.close)
                
        self.layout.addWidget(person_label_rank)
        self.layout.addWidget(person_label_sec_name)
        self.layout.addWidget(person_label_name)
        self.layout.addWidget(line)
        self.layout.addWidget(QLabel("Календарні періоди"))
        self.layout.addWidget(self.cal_table)
        self.layout.addWidget(QLabel("Пільгові періоди"))
        self.layout.addWidget(self.pref_table)
        self.layout.addLayout(button_layout)
        self.layout.addWidget(close_button)

        self.cal_table.itemSelectionChanged.connect(self.clear_pref_selection)
        self.pref_table.itemSelectionChanged.connect(self.clear_cal_selection)

        self.cal_periods_data = []
        self.pref_periods_data = []

        self.load_periods()

    def load_periods(self):
        conn = sqlite3.connect(self.current_db)
        curs = conn.cursor()

        self.cal_periods_data.clear()
        self.pref_periods_data.clear()

        curs.execute("SELECT start_date, end_date FROM service_periods WHERE person_id = ?", (self.person_id,))
        periods = curs.fetchall()
        self.cal_table.setRowCount(len(periods))
        for row_idx, (start, end) in enumerate(periods):
            self.cal_periods_data.append((start, end))
            for col_idx, value in enumerate([start, end]):
                item = QTableWidgetItem(self.format_date(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.cal_table.setItem(row_idx, col_idx, item)
            
        curs.execute("SELECT start_date, end_date, preference_type FROM preferenced_periods WHERE person_id = ?", (self.person_id,))
        pref_periods = curs.fetchall()
        self.pref_table.setRowCount(len(pref_periods))
        for row_idx, (start, end, pref) in enumerate(pref_periods):
            self.pref_periods_data.append((start, end, pref))
            for col_idx, value in enumerate([start, end]):
                item = QTableWidgetItem(self.format_date(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.pref_table.setItem(row_idx, col_idx, item)

            pref_display = "Цивільний ВНЗ" if pref == "Навчання у цивільному ВНЗ" else pref
            pref_item = QTableWidgetItem(pref_display)
            pref_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            pref_item.setFlags(pref_item.flags() ^ Qt.ItemFlag.ItemIsEditable)
            self.pref_table.setItem(row_idx, 2, pref_item)

        self.pref_table.resizeRowsToContents()
        self.cal_table.resizeRowsToContents()
            
        conn.close()
    
    def get_selected_period(self):
        selected_items_cal = self.cal_table.selectedItems()
        selected_items_pref = self.pref_table.selectedItems()

        if selected_items_cal and not selected_items_pref:
            return 'cal', self.cal_table.currentRow()
        elif selected_items_pref and not selected_items_cal:
            return 'pref', self.pref_table.currentRow()
        else:
            return None, -1
        
    def clear_cal_selection(self):
        self.cal_table.blockSignals(True)
        self.cal_table.clearSelection()
        self.cal_table.blockSignals(False)

    def clear_pref_selection(self):
        self.pref_table.blockSignals(True)
        self.pref_table.clearSelection()
        self.pref_table.blockSignals(False)

    def edit_selected_period(self):
        ptype, row = self.get_selected_period()
        if row < 0:
            return
        
        if ptype == 'cal':
            old_start, old_end = self.cal_periods_data[row]
                        
            dialog = AddPeriod_Calendar(self, self.person_data, self.person_id)

            d1, m1, y1 = map(int, old_start.split("-")[2]), int(old_start.split("-")[1]), int(old_start.split("-")[0])
            dialog.start_day_input.setCurrentText(str(d1))
            dialog.start_month_input.setCurrentText(str(m1))
            dialog.start_year_input.setCurrentText(str(y1))

            if old_end == "NOW":
                dialog.now_checkbox.setChecked(True)
            else:
                d2, m2, y2 = map(int, old_end.split("-")[2]), int(old_end.split("-")[1]), int(old_end.split("-")[0])
                dialog.end_day_input.setCurrentText(str(d2))
                dialog.end_month_input.setCurrentText(str(m2))
                dialog.end_year_input.setCurrentText(str(y2))

            if dialog.exec():
                new_data = dialog.get_info_data_period_cal()
                conn = sqlite3.connect(self.current_db)
                curs = conn.cursor()
                curs.execute("""
                    UPDATE service_periods SET start_date=?, end_date=?
                    WHERE person_id=? AND start_date=? AND end_date=?
                """, (
                    new_data["start_date"], new_data["end_date"],
                    self.person_id, old_start, old_end
                ))
                conn.commit()
                conn.close()
                self.load_periods()
                
        elif ptype =='pref':
            old_start, old_end, old_pref = self.pref_periods_data[row]
            
            dialog = AddPeriod_Pref(self, self.person_data, self.person_id)

            d1, m1, y1 = map(int, old_start.split("-")[2]), int(old_start.split("-")[1]), int(old_start.split("-")[0])
            dialog.start_day_input.setCurrentText(str(d1))
            dialog.start_month_input.setCurrentText(str(m1))
            dialog.start_year_input.setCurrentText(str(y1))
                        
            if old_end == "NOW":
                dialog.now_checkbox.setChecked(True)
            else:
                d2, m2, y2 = map(int, old_end.split("-")[2]), int(old_end.split("-")[1]), int(old_end.split("-")[0])
                dialog.end_day_input.setCurrentText(str(d2))
                dialog.end_month_input.setCurrentText(str(m2))
                dialog.end_year_input.setCurrentText(str(y2))

            dialog.pref_type_input.setCurrentText(old_pref)
           
            if dialog.exec():
                new_data = dialog.get_info_data_period_pref()
                conn = sqlite3.connect(self.current_db)
                curs = conn.cursor()
                curs.execute("""
                    UPDATE preferenced_periods SET start_date=?, end_date=?, preference_type=?
                    WHERE person_id=? AND start_date=? AND end_date=? AND preference_type=?
                """, (
                    new_data["start_date"], new_data["end_date"], new_data["preference_type"],
                    self.person_id, old_start, old_end, old_pref
                ))
                conn.commit()
                conn.close()
                self.load_periods()
                
    def del_selected_period(self):
        ptype, row = self.get_selected_period()
        if row < 0:
            return
        
        if ptype == 'cal':
            start, end = self.cal_periods_data[row]
            start_disp = self.cal_table.item(row, 0).text()
            end_disp = self.cal_table.item(row, 1).text()
            msg = f"Видалити календарний період:\n{start_disp} - {end_disp}?"
            query = "DELETE FROM service_periods WHERE person_id = ? AND start_date = ? AND end_date = ?"
            params = (self.person_id, start, end)

        else:
            start, end, pref = self.pref_periods_data[row]
            start_disp = self.pref_table.item(row, 0).text()
            end_disp = self.pref_table.item(row, 1).text()
            pref_disp = self.pref_table.item(row, 2).text()
            msg = f"Видалити пільговий період:\n{start_disp} - {end_disp} ({pref_disp})?"
            query = "DELETE FROM preferenced_periods WHERE person_id = ? AND start_date = ? AND end_date = ? AND preference_type = ?"
            params = (self.person_id, start, end, pref)
            
        confirm = QMessageBox.question(
                self,
                "Підтвердження видалення", msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

        if confirm == QMessageBox.StandardButton.Yes:
            conn = sqlite3.connect(self.current_db)
            curs = conn.cursor()
            curs.execute(query, params)
            conn.commit()
            conn.close()
            self.load_periods()

    def format_date(self, date_str):
        if date_str == "NOW":
            return "по т.ч."
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            return "???"

class CreateDatabaseDialog(QDialog):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setWindowTitle("Створити базу даних")
        self.setGeometry(300, 300, 300, 100)
        self.setFixedSize(300, 100)

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.input = QLineEdit()
        self.input.setPlaceholderText("Введіть номер військової частини")
        self.layout.addWidget(self.input)

        button_layout = QHBoxLayout()
        self.ok_button = QPushButton("Створити")
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Скасувати")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)

        self.layout.addLayout(button_layout)

    def get_database_name(self):
        name = self.input.text().strip().replace(" ", "_")
        return f"{name}.db" if name else None

class AboutDialog(QDialog):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setWindowTitle("Про програму")
        self.setGeometry(150, 150, 180, 70)
        self.setFixedSize(180, 70)

        self.layout = QVBoxLayout()
        self.layout.addWidget(QLabel("Підрахунок вислуги років.\nВерсія програми 0.7.8_beta"))
        self.setLayout(self.layout)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainProg()
    window.show()
    sys.exit(app.exec())


